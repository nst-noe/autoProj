import openpyxl
from openpyxl.styles import Alignment
workbook = openpyxl.load_workbook("data.xlsx")
wb = openpyxl.load_workbook("1c_Data.xlsx")
worksheet = workbook.active
worksheet.merge_cells('A1:I1')
worksheet['A1'] = 'КАРТА ДИСЦИПЛИН\n 09.03.03 Прикладная информатика\n Профиль "Корпоративные информационные системы", 2021 год набора, очная форма обучения'
worksheet.column_dimensions['A'].height = 55
worksheet.row_dimensions[1].height = 55
worksheet.column_dimensions['A'].wight = 2000
worksheet.row_dimensions[1].wight = 2000
currentCell = worksheet['A1']
currentCell.alignment = Alignment(horizontal='center')
workbook.save(filename="data.xlsx")
ws = wb.active 
time = 0
key = 2
test = 'random'
line = 0
save = 0
term = 'B'
Name = ws["D2"].value
worksheet["A2"] = "З.Е."
currentCell = worksheet['A2']
currentCell.alignment = Alignment(horizontal='center')
for col in range(3, 33):                                                  ###
    worksheet["A" + str(col)] = col - 2                                     ###
    currentCell = worksheet["A" + str(col)]                                   ###
    currentCell.alignment = Alignment(horizontal='center')######################## Тут нужен жесткий баг
for col in range(ord('B'), ord('J')):                                         ###
    worksheet[chr(col) + str(2)] = str(col - 65) + " семестр"               ###
    currentCell = worksheet[chr(col) + str(2)]                            ###
    currentCell.alignment = Alignment(horizontal='center')      
for i in range(2, ws.max_row + 1):  
    if ws["I" + str(i)].value != None:
        time = time + ws["I" + str(i)].value 
        if ws["D" + str(i)].value != ws["D" + str(i + 1)].value:
            if ws["E" + str(i)].value == ws["E" + str(i + 1)].value:
                test = term + str(key + 1) 
                worksheet[test] = Name
                key += 1
            # elif ws["E" + str(i)].value == ws["E2"].value and i > 25:  #
            #      line += 1
            #      term = 'B'
            #      key = line
            #      test = term + str(key + 1)
            #      worksheet[test] = Name                                #
            else:
                line = key
                key = 2
                term = chr(ord(term) + 1)
                test = term + str(key + 1)
                worksheet[test] = Name
    else:
        time = 0 
        Name = ws["D" + str(i)].value
workbook.save(filename="data.xlsx")


